"""
Batch evaluation harness for the 50-question gold set
For each question in the 'Gold Set' sheet of the evaluation workbook, this script runs the real RAG pipeline
(the same functions the Streamlit app and generate_answer.py use, imported directly so no logic is re-implemented)
and writes the machine-measurable outputs back into the sheet:
"""

import os
import sys
import time
import argparse
import openpyxl
from groq import Groq
#import the live pipeline
from config import GOLD_SET_FILE
from core.generate_answer import get_collection, generate_answer, retrieve_context
from core.prerequisite_check import detect_prerequisite
#Configuration
WORKBOOK_FILE = str(GOLD_SET_FILE)   #resolved from config
SHEET_NAME    = "Gold Set"
FIRST_DATA_ROW = 2   #row 1 is the header



#Column indices (1-based) in the Gold Set sheet
COL_QUESTION = 5    #E  (input)
COL_SYS_ANS  = 8    #H  (output)
COL_LATENCY  = 13   #M  (output)
COL_CONF     = 14   #N  (output)
COL_PREREQ   = 15   #O  (output)
#Appended output columns (added if missing so no existing letter-based
# Results Summary reference is disturbed):
COL_CONTEXT  = 17   #Q  (output) Retrieved Context
COL_SIGNAL1  = 18   #R  (output) Signal 1: band | mean dist
COL_SIGNAL2  = 19   #S  (output) Signal 2: verify score

APPENDED_HEADERS = {
    COL_CONTEXT: "Retrieved Context (chunks)",
    COL_SIGNAL1: "Signal 1 (band | mean dist)",
    COL_SIGNAL2: "Signal 2 (verify 1-3)",
}

SLEEP_BETWEEN_QUESTIONS = 2.0   #seconds between questions
MAX_RETRIES = 4                 # per   question, on API/rate-limit error
BACKOFF_BASE = 5.0              # seconds, grows 5, 10, 20, 40...

#Helpers
def format_context(retrieved_chunks: list[dict]) -> str:
    """Render the retrieved chunks as a readable, labelled block for the expert to read while scoring Faithfulness (M2)
    and Context Relevance (M4). Full chunktext is kept (not truncated) so faithfulness can be judged properly.
    """
    blocks = []
    for i, c in enumerate(retrieved_chunks, start=1):
        meta = c["metadata"]
        header = (f"[Source {i} | Week {meta['week']} | {meta['topic_label']} "
                  f"| {meta['artefact_type']} | dist={round(c['distance'], 3)}]")
        blocks.append(f"{header}\n{c['text']}")
    return "\n\n".join(blocks)




def run_one_question(client, collection, question: str) -> dict:
    """Run the full pipeline for a single question, with retry/backoff.
    Returns a dict of the cell values to write, or raises after MAX_RETRIES.
    """
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            #Retrieve chunk text separately (generate_answer's sources omit text)
            retrieved = retrieve_context(collection, question)
            #Full pipeline: answer + compound confidence + latency
            result = generate_answer(client, question, collection)
            #Trigger detection only (non-interactive; no MCQ prompt)
            prereq_fires = detect_prerequisite(question) is not None

            return {
                "answer":   result["answer"],
                "latency":  result["latency_seconds"],
                "conf":     result["confidence"],
                "prereq":   "Y" if prereq_fires else "N",
                "context":  format_context(retrieved),
                "signal1":  f"{result['retrieval_band']} | {result['mean_distance']}",
                "signal2":  result["verify_score"],
            }
        except Exception as e:           #rate limit, transient API error, etc.
            last_err = e
            wait = BACKOFF_BASE * (2 ** (attempt - 1))
            print(f"      ! attempt {attempt}/{MAX_RETRIES} failed: {e}")
            if attempt < MAX_RETRIES:
                print(f"      … backing off {wait:.0f}s and retrying")
                time.sleep(wait)
    #all retries used, write the error into the sheet rather than crashing
    raise RuntimeError(str(last_err))


def ensure_appended_headers(ws) -> None:
    """Write headers for the appended columns if they aren't there yet."""
    for col, title in APPENDED_HEADERS.items():
        if ws.cell(row=1, column=col).value != title:
            ws.cell(row=1, column=col, value=title)

#Main
def main() -> None:
    parser = argparse.ArgumentParser(description="Batch-run the gold set through the RAG pipeline.")
    parser.add_argument("--file", default=WORKBOOK_FILE, help="path to the evaluation .xlsx")
    parser.add_argument("--force", action="store_true", help="redo rows even if an answer is already present")
    args = parser.parse_args()

    api_key = os.environ.get("GROQ_API_KEY", "").strip()
    if not api_key:
        sys.exit('\n[ERROR] GROQ_API_KEY not set.\n  export GROQ_API_KEY="your-key-here"\n')

    if not os.path.isfile(args.file):
        sys.exit(f"\n[ERROR] Workbook not found: {args.file}\n")

    print("=" * 70)
    print(" Gold set batch evaluation")
    print("=" * 70)

    print("\nConnecting to knowledge base…")
    collection = get_collection()
    print(f"Connected. Collection contains {collection.count()} chunks.")
    client = Groq(api_key=api_key)

    wb = openpyxl.load_workbook(args.file)
    ws = wb[SHEET_NAME]
    ensure_appended_headers(ws)

    total = done = skipped = failed = 0

    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        question = ws.cell(row=row, column=COL_QUESTION).value
        if not question or not str(question).strip():
            continue
        total += 1

        already = ws.cell(row=row, column=COL_SYS_ANS).value
        if already and not args.force:
            skipped += 1
            continue

        qnum = ws.cell(row=row, column=1).value
        print(f"\n[{row-1:>2}/50] Q{qnum}: {str(question)[:60]}…")

        try:
            out = run_one_question(client, collection, str(question))
            ws.cell(row=row, column=COL_SYS_ANS, value=out["answer"])
            ws.cell(row=row, column=COL_LATENCY, value=out["latency"])
            ws.cell(row=row, column=COL_CONF,    value=out["conf"])
            ws.cell(row=row, column=COL_PREREQ,  value=out["prereq"])
            ws.cell(row=row, column=COL_CONTEXT, value=out["context"])
            ws.cell(row=row, column=COL_SIGNAL1, value=out["signal1"])
            ws.cell(row=row, column=COL_SIGNAL2, value=out["signal2"])
            done += 1
            print(f"      ✓ conf={out['conf']}  latency={out['latency']}s  "
                  f"prereq={out['prereq']}  sig1=({out['signal1']})  sig2={out['signal2']}")
        except Exception as e:
            ws.cell(row=row, column=COL_SYS_ANS, value=f"[ERROR] {e}")
            failed += 1
            print(f"      ✗ FAILED after retries: {e}")

        # Save after every question so nothing is lost on a later abort
        wb.save(args.file)
        time.sleep(SLEEP_BETWEEN_QUESTIONS)

    print("\n" + "-" * 70)
    print(f"Done. questions={total}  written={done}  skipped(existing)={skipped}  failed={failed}")
    print(f"Saved: {args.file}")
    print("\nNext: score columns I-L (M1-M4) by reading each answer against")
    print("'Must Include' (M1) and the Retrieved Context column (M2, M4). The")
    print("Results Summary sheet then computes means, difficulty/week breakdowns,")
    print("confidence calibration, and prerequisite precision/recall automatically.")




if __name__ == "__main__":
    main()